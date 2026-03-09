"""
API для работы с заявками (новая структура)
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import datetime
from typing import Optional, List
import os
import uuid
import shutil

from .db import get_db
from .auth import get_current_user

router = APIRouter(prefix="/api/tasks", tags=["tasks"])

# Директория для загрузки фото
UPLOAD_DIR = "/app/uploads/photos"
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ==================== СПРАВОЧНИК ВИДОВ РАБОТ ====================

@router.get("/work-types")
def list_work_types(
    active_only: bool = True,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Список видов работ"""
    query = "SELECT id, name, description, is_active, created_at FROM work_types"
    if active_only:
        query += " WHERE is_active = TRUE"
    query += " ORDER BY name"
    
    result = db.execute(text(query))
    return [{
        "id": row[0],
        "name": row[1],
        "description": row[2],
        "is_active": row[3],
        "created_at": str(row[4])
    } for row in result]


@router.post("/work-types")
def create_work_type(
    name: str,
    description: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Создать вид работ (только admin)"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = db.execute(
        text("""
            INSERT INTO work_types (name, description)
            VALUES (:name, :description)
            RETURNING id, name, description, is_active, created_at
        """),
        {"name": name, "description": description}
    ).first()
    db.commit()
    
    return {
        "id": result[0],
        "name": result[1],
        "description": result[2],
        "is_active": result[3],
        "created_at": str(result[4])
    }


# ==================== ЗАЯВКИ ====================

@router.get("")
def list_tasks(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Список заявок"""
    # Монтажники видят только свои заявки
    if current_user["role"] == "technician":
        user_id = current_user["id"]
    
    query = """
        SELECT 
            tr.id,
            tr.user_id,
            u.username,
            tr.task_number,
            tr.account_number,
            tr.subscriber_name,
            tr.city,
            tr.address,
            tr.priority,
            tr.status,
            tr.notes,
            tr.created_at,
            tr.completed_at
        FROM task_requests tr
        JOIN users u ON tr.user_id = u.id
        WHERE 1=1
    """
    params = {}
    
    if user_id:
        query += " AND tr.user_id = :user_id"
        params["user_id"] = user_id
    
    if status:
        query += " AND tr.status = :status"
        params["status"] = status
    
    if date_from:
        query += " AND tr.created_at >= :date_from"
        params["date_from"] = date_from
    
    if date_to:
        query += " AND tr.created_at <= :date_to"
        params["date_to"] = date_to
    
    query += " ORDER BY tr.created_at DESC"
    
    result = db.execute(text(query), params)
    tasks = []
    
    for row in result:
        task = {
            "id": row[0],
            "user_id": row[1],
            "username": row[2],
            "task_number": row[3],
            "account_number": row[4],
            "subscriber_name": row[5],
            "city": row[6],
            "address": row[7],
            "priority": row[8],
            "status": row[9],
            "notes": row[10],
            "created_at": str(row[11]),
            "completed_at": str(row[12]) if row[12] else None,
            "work_items": [],
            "photos": []
        }
        
        # Получаем виды работ
        work_items_result = db.execute(
            text("""
                SELECT twi.id, twi.work_type_id, wt.name as work_type_name, twi.notes
                FROM task_work_items twi
                JOIN work_types wt ON twi.work_type_id = wt.id
                WHERE twi.task_request_id = :task_id
            """),
            {"task_id": task["id"]}
        )
        
        for wi_row in work_items_result:
            work_item = {
                "id": wi_row[0],
                "work_type_id": wi_row[1],
                "work_type_name": wi_row[2],
                "notes": wi_row[3],
                "equipment_items": []
            }
            
            # Получаем оборудование/материалы для вида работ
            equip_result = db.execute(
                text("""
                    SELECT 
                        tei.id,
                        tei.equipment_id,
                        e.name as equipment_name,
                        e.material_number as equipment_material_number,
                        tei.serial_number_id,
                        sn.serial_number,
                        tei.material_id,
                        m.name as material_name,
                        m.material_number as material_material_number,
                        tei.quantity
                    FROM task_equipment_items tei
                    LEFT JOIN equipment e ON tei.equipment_id = e.id
                    LEFT JOIN serial_numbers sn ON tei.serial_number_id = sn.id
                    LEFT JOIN materials m ON tei.material_id = m.id
                    WHERE tei.task_work_item_id = :work_item_id
                """),
                {"work_item_id": work_item["id"]}
            )
            
            for eq_row in equip_result:
                work_item["equipment_items"].append({
                    "id": eq_row[0],
                    "equipment_id": eq_row[1],
                    "equipment_name": eq_row[2],
                    "equipment_material_number": eq_row[3],
                    "serial_number_id": eq_row[4],
                    "serial_number": eq_row[5],
                    "material_id": eq_row[6],
                    "material_name": eq_row[7],
                    "material_material_number": eq_row[8],
                    "quantity": eq_row[9]
                })
            
            task["work_items"].append(work_item)
        
        # Получаем фото (без содержимого, только метаданные)
        photos_result = db.execute(
            text("""
                SELECT id, file_name, description, work_type_id, created_at
                FROM task_photos
                WHERE task_request_id = :task_id
            """),
            {"task_id": task["id"]}
        )
        
        for p_row in photos_result:
            task["photos"].append({
                "id": p_row[0],
                "file_name": p_row[1],
                "description": p_row[2],
                "work_type_id": p_row[3],
                "created_at": str(p_row[4])
            })
        
        tasks.append(task)
    
    return tasks


@router.get("/{task_id}")
def get_task(
    task_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Получить заявку по ID"""
    result = db.execute(
        text("""
            SELECT 
                tr.id, tr.user_id, u.username,
                tr.task_number, tr.account_number, tr.subscriber_name,
                tr.city, tr.address, tr.priority, tr.status, tr.notes,
                tr.created_at, tr.completed_at
            FROM task_requests tr
            JOIN users u ON tr.user_id = u.id
            WHERE tr.id = :id
        """),
        {"id": task_id}
    ).first()
    
    if not result:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Монтажник может смотреть только свои заявки
    if user["role"] == "technician" and result[1] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    task = {
        "id": result[0],
        "user_id": result[1],
        "username": result[2],
        "task_number": result[3],
        "account_number": result[4],
        "subscriber_name": result[5],
        "city": result[6],
        "address": result[7],
        "priority": result[8],
        "status": result[9],
        "notes": result[10],
        "created_at": str(result[11]),
        "completed_at": str(result[12]) if result[12] else None,
        "work_items": [],
        "photos": []
    }
    
    # Получаем виды работ и оборудование (так же как в list_tasks)
    work_items_result = db.execute(
        text("""
            SELECT twi.id, twi.work_type_id, wt.name as work_type_name, twi.notes
            FROM task_work_items twi
            JOIN work_types wt ON twi.work_type_id = wt.id
            WHERE twi.task_request_id = :task_id
        """),
        {"task_id": task["id"]}
    )
    
    for wi_row in work_items_result:
        work_item = {
            "id": wi_row[0],
            "work_type_id": wi_row[1],
            "work_type_name": wi_row[2],
            "notes": wi_row[3],
            "equipment_items": []
        }
        
        equip_result = db.execute(
            text("""
                SELECT 
                    tei.id,
                    tei.equipment_id, e.name as equipment_name,
                    tei.serial_number_id, sn.serial_number,
                    tei.material_id, m.name as material_name,
                    tei.quantity
                FROM task_equipment_items tei
                LEFT JOIN equipment e ON tei.equipment_id = e.id
                LEFT JOIN serial_numbers sn ON tei.serial_number_id = sn.id
                LEFT JOIN materials m ON tei.material_id = m.id
                WHERE tei.task_work_item_id = :work_item_id
            """),
            {"work_item_id": work_item["id"]}
        )
        
        for eq_row in equip_result:
            work_item["equipment_items"].append({
                "id": eq_row[0],
                "equipment_id": eq_row[1],
                "equipment_name": eq_row[2],
                "serial_number_id": eq_row[3],
                "serial_number": eq_row[4],
                "material_id": eq_row[5],
                "material_name": eq_row[6],
                "quantity": eq_row[7]
            })
        
        task["work_items"].append(work_item)
    
    return task


@router.post("")
def create_task(
    task_number: Optional[str] = Form(None),
    account_number: Optional[str] = Form(None),
    subscriber_name: Optional[str] = Form(None),
    city: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    priority: str = Form("Обычный"),
    notes: Optional[str] = Form(None),
    work_items: str = Form(...),  # JSON строка с видами работ и оборудованием
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Создать заявку.
    work_items - JSON строка вида:
    [
        {
            "work_type_id": 1,
            "notes": "примечание",
            "equipment_items": [
                {"equipment_id": 1, "serial_number_id": 5, "quantity": 1},
                {"material_id": 3, "quantity": 10}
            ]
        }
    ]
    """
    import json
    
    try:
        work_items_data = json.loads(work_items)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid work_items JSON")
    
    if not work_items_data:
        raise HTTPException(status_code=400, detail="At least one work item is required")
    
    # Получаем склад пользователя
    user_warehouse = db.execute(
        text("SELECT warehouse_id FROM users WHERE id = :id"),
        {"id": user["id"]}
    ).first()
    
    warehouse_id = user_warehouse[0] if user_warehouse else None
    
    try:
        # Создаем заявку
        task_result = db.execute(
            text("""
                INSERT INTO task_requests 
                (user_id, task_number, account_number, subscriber_name, city, address, priority, notes, status)
                VALUES (:user_id, :task_number, :account_number, :subscriber_name, :city, :address, :priority, :notes, 'new')
                RETURNING id
            """),
            {
                "user_id": user["id"],
                "task_number": task_number,
                "account_number": account_number,
                "subscriber_name": subscriber_name,
                "city": city,
                "address": address,
                "priority": priority,
                "notes": notes
            }
        ).first()
        
        task_id = task_result[0]
        
        # Создаем виды работ и списываем оборудование/материалы
        for work_item in work_items_data:
            work_type_id = work_item.get("work_type_id")
            if not work_type_id:
                continue
            
            # Создаем вид работ
            work_item_result = db.execute(
                text("""
                    INSERT INTO task_work_items (task_request_id, work_type_id, notes)
                    VALUES (:task_id, :work_type_id, :notes)
                    RETURNING id
                """),
                {
                    "task_id": task_id,
                    "work_type_id": work_type_id,
                    "notes": work_item.get("notes")
                }
            ).first()
            
            work_item_id = work_item_result[0]
            
            # Обрабатываем оборудование/материалы
            equipment_items = work_item.get("equipment_items", [])
            for item in equipment_items:
                equipment_id = item.get("equipment_id")
                serial_number_id = item.get("serial_number_id")
                material_id = item.get("material_id")
                quantity = item.get("quantity", 1)
                
                # Создаем запись в task_equipment_items
                db.execute(
                    text("""
                        INSERT INTO task_equipment_items 
                        (task_work_item_id, equipment_id, serial_number_id, material_id, quantity)
                        VALUES (:work_item_id, :equipment_id, :serial_number_id, :material_id, :quantity)
                    """),
                    {
                        "work_item_id": work_item_id,
                        "equipment_id": equipment_id,
                        "serial_number_id": serial_number_id,
                        "material_id": material_id,
                        "quantity": quantity
                    }
                )
                
                # Списываем оборудование (помечаем серийный номер как written_off)
                if serial_number_id and warehouse_id:
                    # Проверяем что серийный номер на складе пользователя
                    sn = db.execute(
                        text("SELECT status, warehouse_id FROM serial_numbers WHERE id = :id"),
                        {"id": serial_number_id}
                    ).first()
                    
                    if sn and sn[1] == warehouse_id and sn[0] == 'available':
                        db.execute(
                            text("UPDATE serial_numbers SET status = 'written_off' WHERE id = :id"),
                            {"id": serial_number_id}
                        )
                        
                        # Запись в историю транзакций
                        db.execute(
                            text("""
                                INSERT INTO inventory_transactions 
                                (equipment_id, serial_number_id, from_warehouse_id, to_warehouse_id, 
                                 quantity, transaction_type, notes, created_by)
                                VALUES (:equipment_id, :sn_id, :warehouse_id, NULL, 1, 'write_off', :notes, :user_id)
                            """),
                            {
                                "equipment_id": equipment_id,
                                "sn_id": serial_number_id,
                                "warehouse_id": warehouse_id,
                                "notes": f"Списание через заявку #{task_id}",
                                "user_id": user["id"]
                            }
                        )
                
                # Списываем материал
                if material_id and warehouse_id:
                    # Проверяем наличие
                    stock = db.execute(
                        text("""
                            SELECT quantity FROM material_stock 
                            WHERE material_id = :material_id AND warehouse_id = :warehouse_id
                        """),
                        {"material_id": material_id, "warehouse_id": warehouse_id}
                    ).first()
                    
                    if stock and stock[0] >= quantity:
                        db.execute(
                            text("""
                                UPDATE material_stock 
                                SET quantity = quantity - :qty 
                                WHERE material_id = :material_id AND warehouse_id = :warehouse_id
                            """),
                            {"qty": quantity, "material_id": material_id, "warehouse_id": warehouse_id}
                        )
                        
                        # Запись в историю
                        db.execute(
                            text("""
                                INSERT INTO material_transactions 
                                (material_id, from_warehouse_id, to_warehouse_id, quantity, transaction_type, notes, created_by)
                                VALUES (:material_id, :warehouse_id, NULL, :qty, 'write_off', :notes, :user_id)
                            """),
                            {
                                "material_id": material_id,
                                "warehouse_id": warehouse_id,
                                "qty": quantity,
                                "notes": f"Списание через заявку #{task_id}",
                                "user_id": user["id"]
                            }
                        )
        
        db.commit()
        
        return {
            "id": task_id,
            "status": "created",
            "message": "Task created successfully"
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating task: {str(e)}")


@router.put("/{task_id}")
def update_task(
    task_id: int,
    status: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Обновить заявку"""
    # Проверяем доступ
    task = db.execute(
        text("SELECT user_id FROM task_requests WHERE id = :id"),
        {"id": task_id}
    ).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if user["role"] == "technician" and task[0] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    updates = []
    params = {"id": task_id}
    
    if status:
        updates.append("status = :status")
        params["status"] = status
        
        if status == "completed":
            updates.append("completed_at = NOW()")
    
    if notes is not None:
        updates.append("notes = :notes")
        params["notes"] = notes
    
    if not updates:
        return {"message": "No updates"}
    
    updates.append("updated_at = NOW()")
    
    db.execute(
        text(f"UPDATE task_requests SET {', '.join(updates)} WHERE id = :id"),
        params
    )
    db.commit()
    
    return {"updated": True, "id": task_id}


@router.delete("/{task_id}")
def delete_task(
    task_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Удалить заявку (только admin)"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Удаляем (каскадно удалит связанные записи)
    result = db.execute(
        text("DELETE FROM task_requests WHERE id = :id RETURNING id"),
        {"id": task_id}
    ).first()
    
    if not result:
        raise HTTPException(status_code=404, detail="Task not found")
    
    db.commit()
    return {"deleted": True, "id": task_id}


# ==================== ЗАГРУЗКА ФОТО ====================

@router.post("/{task_id}/photos")
async def upload_photo(
    task_id: int,
    file: UploadFile = File(...),
    work_type_id: Optional[int] = Form(None),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Загрузить фото к заявке"""
    # Проверяем доступ
    task = db.execute(
        text("SELECT user_id FROM task_requests WHERE id = :id"),
        {"id": task_id}
    ).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if user["role"] == "technician" and task[0] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Генерируем уникальное имя файла
    file_ext = os.path.splitext(file.filename)[1] if file.filename else ".jpg"
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)
    
    # Сохраняем файл
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Записываем в БД
    result = db.execute(
        text("""
            INSERT INTO task_photos (task_request_id, work_type_id, file_path, file_name, description, uploaded_by)
            VALUES (:task_id, :work_type_id, :file_path, :file_name, :description, :user_id)
            RETURNING id
        """),
        {
            "task_id": task_id,
            "work_type_id": work_type_id,
            "file_path": file_path,
            "file_name": file.filename,
            "description": description,
            "user_id": user["id"]
        }
    ).first()
    
    db.commit()
    
    return {
        "id": result[0],
        "file_name": file.filename,
        "message": "Photo uploaded successfully"
    }


@router.get("/{task_id}/photos/{photo_id}")
def get_photo(
    task_id: int,
    photo_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Получить фото"""
    # Проверяем доступ
    task = db.execute(
        text("SELECT user_id FROM task_requests WHERE id = :id"),
        {"id": task_id}
    ).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if user["role"] == "technician" and task[0] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Получаем информацию о фото
    photo = db.execute(
        text("SELECT file_path, file_name FROM task_photos WHERE id = :id AND task_request_id = :task_id"),
        {"id": photo_id, "task_id": task_id}
    ).first()
    
    if not photo:
        raise HTTPException(status_code=404, detail="Photo not found")
    
    if not os.path.exists(photo[0]):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        photo[0],
        filename=photo[1],
        media_type="image/jpeg"
    )


@router.delete("/{task_id}/photos/{photo_id}")
def delete_photo(
    task_id: int,
    photo_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Удалить фото"""
    # Проверяем доступ
    task = db.execute(
        text("SELECT user_id FROM task_requests WHERE id = :id"),
        {"id": task_id}
    ).first()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if user["role"] == "technician" and task[0] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Получаем путь к файлу
    photo = db.execute(
        text("SELECT file_path FROM task_photos WHERE id = :id AND task_request_id = :task_id"),
        {"id": photo_id, "task_id": task_id}
    ).first()
    
    if not photo:
        raise HTTPException(status_code=404, detail="Photo not found")
    
    # Удаляем файл
    if os.path.exists(photo[0]):
        os.remove(photo[0])
    
    # Удаляем запись из БД
    db.execute(
        text("DELETE FROM task_photos WHERE id = :id"),
        {"id": photo_id}
    )
    db.commit()
    
    return {"deleted": True, "id": photo_id}


# ==================== ОСТАТКИ НА СКЛАДЕ ====================

@router.get("/my-stock")
def get_my_stock(
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Получить остатки на складе текущего пользователя"""
    # Получаем склад пользователя
    warehouse = db.execute(
        text("SELECT warehouse_id FROM users WHERE id = :id"),
        {"id": user["id"]}
    ).first()
    
    if not warehouse or not warehouse[0]:
        return {"equipment": [], "materials": [], "warehouse_id": None}
    
    warehouse_id = warehouse[0]
    
    # Получаем оборудование с серийными номерами
    equipment_query = text("""
        SELECT 
            e.id as equipment_id,
            e.material_number,
            e.name as equipment_name,
            e.category,
            e.unit,
            sn.id as serial_id,
            sn.serial_number,
            sn.status
        FROM serial_numbers sn
        JOIN equipment e ON sn.equipment_id = e.id
        WHERE sn.warehouse_id = :warehouse_id 
          AND sn.status = 'available'
        ORDER BY e.name, sn.serial_number
    """)
    
    equipment_result = db.execute(equipment_query, {"warehouse_id": warehouse_id})
    equipment_list = [{
        "equipment_id": row[0],
        "material_number": row[1],
        "name": row[2],
        "category": row[3],
        "unit": row[4],
        "serial_id": row[5],
        "serial_number": row[6],
        "status": row[7]
    } for row in equipment_result]
    
    # Получаем материалы
    materials_query = text("""
        SELECT 
            m.id as material_id,
            m.material_number,
            m.name as material_name,
            m.category,
            m.unit,
            ms.quantity
        FROM material_stock ms
        JOIN materials m ON ms.material_id = m.id
        WHERE ms.warehouse_id = :warehouse_id AND ms.quantity > 0
        ORDER BY m.name
    """)
    
    materials_result = db.execute(materials_query, {"warehouse_id": warehouse_id})
    materials_list = [{
        "material_id": row[0],
        "material_number": row[1],
        "name": row[2],
        "category": row[3],
        "unit": row[4],
        "quantity": row[5]
    } for row in materials_result]
    
    return {
        "warehouse_id": warehouse_id,
        "equipment": equipment_list,
        "materials": materials_list
    }


# ==================== ЭКСПОРТ ====================

@router.get("/export/csv")
def export_tasks_csv(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Экспорт заявок в CSV"""
    import io
    import csv
    
    # Монтажники могут экспортировать только свои заявки
    if current_user["role"] == "technician":
        user_id = current_user["id"]
    
    tasks = list_tasks(status, date_from, date_to, user_id, db, current_user)
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Заголовок
    headers = [
        "ID", "№ Таска", "Лицевой счет", "ФИО абонента", "Город", "Адрес",
        "Приоритет", "Вид работ", "Оборудование", "Серийный номер",
        "Материал", "Количество", "Статус", "Создан"
    ]
    writer.writerow(headers)
    
    # Данные
    status_labels = {
        "new": "Новая",
        "in_progress": "В работе",
        "completed": "Завершена",
        "cancelled": "Отменена"
    }
    
    for task in tasks:
        for work_item in task["work_items"]:
            if work_item["equipment_items"]:
                for eq in work_item["equipment_items"]:
                    row = [
                        task["id"],
                        task["task_number"] or "",
                        task["account_number"] or "",
                        task["subscriber_name"] or "",
                        task["city"] or "",
                        task["address"] or "",
                        task["priority"] or "",
                        work_item["work_type_name"],
                        eq.get("equipment_name") or "",
                        eq.get("serial_number") or "",
                        eq.get("material_name") or "",
                        eq.get("quantity", 1),
                        status_labels.get(task["status"], task["status"]),
                        task["created_at"][:19] if task["created_at"] else ""
                    ]
                    writer.writerow(row)
            else:
                row = [
                    task["id"],
                    task["task_number"] or "",
                    task["account_number"] or "",
                    task["subscriber_name"] or "",
                    task["city"] or "",
                    task["address"] or "",
                    task["priority"] or "",
                    work_item["work_type_name"],
                    "", "", "", "",
                    status_labels.get(task["status"], task["status"]),
                    task["created_at"][:19] if task["created_at"] else ""
                ]
                writer.writerow(row)
    
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=tasks.csv"}
    )


@router.get("/export/excel")
def export_tasks_excel(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Экспорт заявок в Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not installed")
    
    import io
    
    # Монтажники могут экспортировать только свои заявки
    if current_user["role"] == "technician":
        user_id = current_user["id"]
    
    tasks = list_tasks(status, date_from, date_to, user_id, db, current_user)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = [
        "ID", "№ Таска", "Лицевой счет", "ФИО абонента", "Город", "Адрес",
        "Приоритет", "Вид работ", "Оборудование", "Серийный номер",
        "Материал", "Количество", "Статус", "Создан"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    status_labels = {
        "new": "Новая",
        "in_progress": "В работе",
        "completed": "Завершена",
        "cancelled": "Отменена"
    }
    
    row_idx = 2
    for task in tasks:
        for work_item in task["work_items"]:
            if work_item["equipment_items"]:
                for eq in work_item["equipment_items"]:
                    row_data = [
                        task["id"],
                        task["task_number"] or "",
                        task["account_number"] or "",
                        task["subscriber_name"] or "",
                        task["city"] or "",
                        task["address"] or "",
                        task["priority"] or "",
                        work_item["work_type_name"],
                        eq.get("equipment_name") or "",
                        eq.get("serial_number") or "",
                        eq.get("material_name") or "",
                        eq.get("quantity", 1),
                        status_labels.get(task["status"], task["status"]),
                        task["created_at"][:10] if task["created_at"] else ""
                    ]
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = border
                    row_idx += 1
            else:
                row_data = [
                    task["id"],
                    task["task_number"] or "",
                    task["account_number"] or "",
                    task["subscriber_name"] or "",
                    task["city"] or "",
                    task["address"] or "",
                    task["priority"] or "",
                    work_item["work_type_name"],
                    "", "", "", "",
                    status_labels.get(task["status"], task["status"]),
                    task["created_at"][:10] if task["created_at"] else ""
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                row_idx += 1
    
    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tasks.xlsx"}
    )