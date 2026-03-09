"""
Модуль отчетности монтажников
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import datetime, timedelta
from typing import Optional
import io
import csv

from .db import get_db
from .auth import get_current_user

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ==================== СПРАВОЧНИК ОБЪЕКТОВ ====================

@router.get("/objects")
def list_work_objects(
    active_only: bool = True,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Список объектов для выполнения работ"""
    query = "SELECT id, name, address, description, is_active, created_at FROM work_objects"
    if active_only:
        query += " WHERE is_active = TRUE"
    query += " ORDER BY name"
    
    result = db.execute(text(query))
    return [{
        "id": row[0],
        "name": row[1],
        "address": row[2],
        "description": row[3],
        "is_active": row[4],
        "created_at": str(row[5])
    } for row in result]


@router.post("/objects")
def create_work_object(
    name: str,
    address: Optional[str] = None,
    description: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Создать новый объект (доступно admin и technician)"""
    if user["role"] not in ["admin", "technician"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = db.execute(
        text("""
            INSERT INTO work_objects (name, address, description)
            VALUES (:name, :address, :description)
            RETURNING id, name, address, description, is_active, created_at
        """),
        {"name": name, "address": address, "description": description}
    ).first()
    db.commit()
    
    return {
        "id": result[0],
        "name": result[1],
        "address": result[2],
        "description": result[3],
        "is_active": result[4],
        "created_at": str(result[5])
    }


@router.put("/objects/{object_id}")
def update_work_object(
    object_id: int,
    name: Optional[str] = None,
    address: Optional[str] = None,
    description: Optional[str] = None,
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Обновить объект (только admin)"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Получаем текущие данные
    current = db.execute(
        text("SELECT name, address, description, is_active FROM work_objects WHERE id = :id"),
        {"id": object_id}
    ).first()
    
    if not current:
        raise HTTPException(status_code=404, detail="Object not found")
    
    # Обновляем
    result = db.execute(
        text("""
            UPDATE work_objects 
            SET name = COALESCE(:name, name),
                address = COALESCE(:address, address),
                description = COALESCE(:description, description),
                is_active = COALESCE(:is_active, is_active)
            WHERE id = :id
            RETURNING id, name, address, description, is_active, created_at
        """),
        {
            "id": object_id,
            "name": name,
            "address": address,
            "description": description,
            "is_active": is_active
        }
    ).first()
    db.commit()
    
    return {
        "id": result[0],
        "name": result[1],
        "address": result[2],
        "description": result[3],
        "is_active": result[4],
        "created_at": str(result[5])
    }


@router.delete("/objects/{object_id}")
def delete_work_object(
    object_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Удалить объект (только admin)"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = db.execute(
        text("DELETE FROM work_objects WHERE id = :id RETURNING id"),
        {"id": object_id}
    ).first()
    
    if not result:
        raise HTTPException(status_code=404, detail="Object not found")
    
    db.commit()
    return {"deleted": True, "id": object_id}


# ==================== ОСТАТКИ НА СКЛАДЕ МОНТАЖНИКА ====================

@router.get("/my-stock")
def get_my_stock(
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Получить остатки на складе текущего пользователя"""
    # Получаем склад пользователя
    warehouse = db.execute(
        text("SELECT warehouse_id FROM users WHERE id = :id"),
        {"id": user["id"]}
    ).first()
    
    if not warehouse or not warehouse[0]:
        return {"equipment": [], "materials": [], "warehouse_id": None}
    
    warehouse_id = warehouse[0]
    
    # Получаем оборудование с серийными номерами на складе монтажника
    equipment_query = text("""
        SELECT 
            e.id as equipment_id,
            e.material_number,
            e.name as equipment_name,
            e.category,
            e.unit,
            sn.id as serial_id,
            sn.serial_number,
            sn.status
        FROM serial_numbers sn
        JOIN equipment e ON sn.equipment_id = e.id
        WHERE sn.warehouse_id = :warehouse_id 
          AND sn.status = 'available'
        ORDER BY e.name, sn.serial_number
    """)
    
    equipment_result = db.execute(equipment_query, {"warehouse_id": warehouse_id})
    equipment_list = [{
        "equipment_id": row[0],
        "material_number": row[1],
        "name": row[2],
        "category": row[3],
        "unit": row[4],
        "serial_id": row[5],
        "serial_number": row[6],
        "status": row[7]
    } for row in equipment_result]
    
    # Получаем материалы на складе монтажника
    materials_query = text("""
        SELECT 
            m.id as material_id,
            m.material_number,
            m.name as material_name,
            m.category,
            m.unit,
            ms.quantity
        FROM material_stock ms
        JOIN materials m ON ms.material_id = m.id
        WHERE ms.warehouse_id = :warehouse_id AND ms.quantity > 0
        ORDER BY m.name
    """)
    
    materials_result = db.execute(materials_query, {"warehouse_id": warehouse_id})
    materials_list = [{
        "material_id": row[0],
        "material_number": row[1],
        "name": row[2],
        "category": row[3],
        "unit": row[4],
        "quantity": row[5]
    } for row in materials_result]
    
    return {
        "warehouse_id": warehouse_id,
        "equipment": equipment_list,
        "materials": materials_list
    }


# ==================== ОТЧЕТЫ ====================

@router.get("")
def list_reports(
    user_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    work_object_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """
    Список отчетов.
    Монтажник видит только свои отчеты.
    Админ видит все отчеты.
    """
    # Монтажники видят только свои отчеты
    if current_user["role"] == "technician":
        user_id = current_user["id"]
    
    query = """
        SELECT 
            wr.id,
            wr.user_id,
            u.username,
            wr.work_date,
            wr.work_object_id,
            wr.object_name,
            wr.object_address,
            wr.equipment_id,
            e.name as equipment_name,
            e.material_number as equipment_material_number,
            wr.serial_number_id,
            sn.serial_number,
            wr.material_id,
            m.name as material_name,
            m.material_number as material_material_number,
            wr.quantity,
            wr.notes,
            wr.status,
            wr.created_at,
            wr.cancelled_at,
            wr.cancel_reason,
            cu.username as cancelled_by_name
        FROM work_reports wr
        JOIN users u ON wr.user_id = u.id
        LEFT JOIN equipment e ON wr.equipment_id = e.id
        LEFT JOIN serial_numbers sn ON wr.serial_number_id = sn.id
        LEFT JOIN materials m ON wr.material_id = m.id
        LEFT JOIN users cu ON wr.cancelled_by = cu.id
        WHERE 1=1
    """
    params = {}
    
    if user_id:
        query += " AND wr.user_id = :user_id"
        params["user_id"] = user_id
    
    if date_from:
        query += " AND wr.work_date >= :date_from"
        params["date_from"] = date_from
    
    if date_to:
        query += " AND wr.work_date <= :date_to"
        params["date_to"] = date_to
    
    if status:
        query += " AND wr.status = :status"
        params["status"] = status
    
    if work_object_id:
        query += " AND wr.work_object_id = :work_object_id"
        params["work_object_id"] = work_object_id
    
    query += " ORDER BY wr.work_date DESC, wr.created_at DESC"
    
    result = db.execute(text(query), params)
    
    return [{
        "id": row[0],
        "user_id": row[1],
        "username": row[2],
        "work_date": str(row[3]),
        "work_object_id": row[4],
        "object_name": row[5],
        "object_address": row[6],
        "equipment_id": row[7],
        "equipment_name": row[8],
        "equipment_material_number": row[9],
        "serial_number_id": row[10],
        "serial_number": row[11],
        "material_id": row[12],
        "material_name": row[13],
        "material_material_number": row[14],
        "quantity": row[15],
        "notes": row[16],
        "status": row[17],
        "created_at": str(row[18]),
        "cancelled_at": str(row[19]) if row[19] else None,
        "cancel_reason": row[20],
        "cancelled_by_name": row[21]
    } for row in result]


@router.get("/{report_id}")
def get_report(
    report_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Получить детали отчета"""
    result = db.execute(
        text("""
            SELECT 
                wr.id,
                wr.user_id,
                u.username,
                wr.work_date,
                wr.work_object_id,
                wo.name as work_object_name,
                wo.address as work_object_address,
                wr.object_name,
                wr.object_address,
                wr.equipment_id,
                e.name as equipment_name,
                e.material_number as equipment_material_number,
                wr.serial_number_id,
                sn.serial_number,
                wr.material_id,
                m.name as material_name,
                m.material_number as material_material_number,
                wr.quantity,
                wr.notes,
                wr.status,
                wr.created_at,
                wr.cancelled_at,
                wr.cancel_reason,
                cu.username as cancelled_by_name
            FROM work_reports wr
            JOIN users u ON wr.user_id = u.id
            LEFT JOIN work_objects wo ON wr.work_object_id = wo.id
            LEFT JOIN equipment e ON wr.equipment_id = e.id
            LEFT JOIN serial_numbers sn ON wr.serial_number_id = sn.id
            LEFT JOIN materials m ON wr.material_id = m.id
            LEFT JOIN users cu ON wr.cancelled_by = cu.id
            WHERE wr.id = :id
        """),
        {"id": report_id}
    ).first()
    
    if not result:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Монтажник может смотреть только свои отчеты
    if user["role"] == "technician" and result[1] != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return {
        "id": result[0],
        "user_id": result[1],
        "username": result[2],
        "work_date": str(result[3]),
        "work_object_id": result[4],
        "work_object_name": result[5],
        "work_object_address": result[6],
        "object_name": result[7],
        "object_address": result[8],
        "equipment_id": result[9],
        "equipment_name": result[10],
        "equipment_material_number": result[11],
        "serial_number_id": result[12],
        "serial_number": result[13],
        "material_id": result[14],
        "material_name": result[15],
        "material_material_number": result[16],
        "quantity": result[17],
        "notes": result[18],
        "status": result[19],
        "created_at": str(result[20]),
        "cancelled_at": str(result[21]) if result[21] else None,
        "cancel_reason": result[22],
        "cancelled_by_name": result[23]
    }


@router.post("")
def create_report(
    work_date: str,
    work_object_id: Optional[int] = None,
    object_name: Optional[str] = None,
    object_address: Optional[str] = None,
    equipment_id: Optional[int] = None,
    serial_number_id: Optional[int] = None,
    material_id: Optional[int] = None,
    quantity: int = 1,
    notes: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Создать отчет о выполненной работе.
    
    При создании происходит автоматическое списание:
    - Для оборудования: серийный номер помечается как 'written_off'
    - Для материалов: уменьшается количество на складе монтажника
    """
    # Валидация: должно быть указано либо оборудование, либо материал
    if not equipment_id and not material_id:
        raise HTTPException(status_code=400, detail="Either equipment_id or material_id is required")
    
    if equipment_id and material_id:
        raise HTTPException(status_code=400, detail="Cannot specify both equipment and material")
    
    # Получаем склад пользователя
    user_warehouse = db.execute(
        text("SELECT warehouse_id FROM users WHERE id = :id"),
        {"id": user["id"]}
    ).first()
    
    if not user_warehouse or not user_warehouse[0]:
        raise HTTPException(status_code=400, detail="User has no assigned warehouse")
    
    warehouse_id = user_warehouse[0]
    
    # Получаем название объекта если выбран из справочника
    if work_object_id and not object_name:
        obj = db.execute(
            text("SELECT name, address FROM work_objects WHERE id = :id"),
            {"id": work_object_id}
        ).first()
        if obj:
            object_name = obj[0]
            if not object_address:
                object_address = obj[1]
    
    # Начинаем транзакцию
    try:
        if equipment_id:
            # Списание оборудования по серийному номеру
            if not serial_number_id:
                raise HTTPException(status_code=400, detail="serial_number_id is required for equipment")
            
            # Проверяем, что серийный номер на складе пользователя и доступен
            sn = db.execute(
                text("""
                    SELECT id, status, warehouse_id 
                    FROM serial_numbers 
                    WHERE id = :id AND equipment_id = :equipment_id
                """),
                {"id": serial_number_id, "equipment_id": equipment_id}
            ).first()
            
            if not sn:
                raise HTTPException(status_code=404, detail="Serial number not found for this equipment")
            
            if sn[2] != warehouse_id:
                raise HTTPException(status_code=400, detail="Serial number is not on your warehouse")
            
            if sn[1] != 'available':
                raise HTTPException(status_code=400, detail=f"Serial number status is '{sn[1]}', cannot write off")
            
            # Помечаем серийный номер как списанный
            db.execute(
                text("UPDATE serial_numbers SET status = 'written_off' WHERE id = :id"),
                {"id": serial_number_id}
            )
            
            # Создаем запись в истории транзакций
            db.execute(
                text("""
                    INSERT INTO inventory_transactions 
                    (equipment_id, serial_number_id, from_warehouse_id, to_warehouse_id, 
                     quantity, transaction_type, notes, created_by)
                    VALUES (:equipment_id, :sn_id, :warehouse_id, NULL, 1, 'write_off', :notes, :user_id)
                """),
                {
                    "equipment_id": equipment_id,
                    "sn_id": serial_number_id,
                    "warehouse_id": warehouse_id,
                    "notes": f"Списание через отчет: {notes or ''}",
                    "user_id": user["id"]
                }
            )
        
        else:
            # Списание материала
            # Проверяем наличие материала на складе
            stock = db.execute(
                text("""
                    SELECT quantity FROM material_stock 
                    WHERE material_id = :material_id AND warehouse_id = :warehouse_id
                """),
                {"material_id": material_id, "warehouse_id": warehouse_id}
            ).first()
            
            if not stock or stock[0] < quantity:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Insufficient material stock. Available: {stock[0] if stock else 0}"
                )
            
            # Уменьшаем количество
            db.execute(
                text("""
                    UPDATE material_stock 
                    SET quantity = quantity - :qty 
                    WHERE material_id = :material_id AND warehouse_id = :warehouse_id
                """),
                {"qty": quantity, "material_id": material_id, "warehouse_id": warehouse_id}
            )
            
            # Создаем запись в истории транзакций
            db.execute(
                text("""
                    INSERT INTO material_transactions 
                    (material_id, from_warehouse_id, to_warehouse_id, quantity, transaction_type, notes, created_by)
                    VALUES (:material_id, :warehouse_id, NULL, :qty, 'write_off', :notes, :user_id)
                """),
                {
                    "material_id": material_id,
                    "warehouse_id": warehouse_id,
                    "qty": quantity,
                    "notes": f"Списание через отчет: {notes or ''}",
                    "user_id": user["id"]
                }
            )
        
        # Создаем отчет
        result = db.execute(
            text("""
                INSERT INTO work_reports 
                (user_id, work_date, work_object_id, object_name, object_address, 
                 equipment_id, serial_number_id, material_id, quantity, notes, status)
                VALUES (:user_id, :work_date, :work_object_id, :object_name, :object_address,
                        :equipment_id, :serial_number_id, :material_id, :quantity, :notes, 'submitted')
                RETURNING id
            """),
            {
                "user_id": user["id"],
                "work_date": work_date,
                "work_object_id": work_object_id,
                "object_name": object_name,
                "object_address": object_address,
                "equipment_id": equipment_id,
                "serial_number_id": serial_number_id,
                "material_id": material_id,
                "quantity": quantity,
                "notes": notes
            }
        ).first()
        
        db.commit()
        
        return {
            "id": result[0],
            "status": "created",
            "message": "Report created and items written off successfully"
        }
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating report: {str(e)}")


@router.post("/{report_id}/cancel")
def cancel_report(
    report_id: int,
    reason: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Отменить отчет с возвратом материалов/оборудования на склад.
    Доступно только админу.
    """
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can cancel reports")
    
    # Получаем отчет
    report = db.execute(
        text("SELECT * FROM work_reports WHERE id = :id"),
        {"id": report_id}
    ).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    if report[12] == 'cancelled':  # status
        raise HTTPException(status_code=400, detail="Report is already cancelled")
    
    report_dict = {
        "id": report[0],
        "user_id": report[1],
        "equipment_id": report[6],
        "serial_number_id": report[7],
        "material_id": report[8],
        "quantity": report[9]
    }
    
    try:
        # Получаем склад пользователя
        user_warehouse = db.execute(
            text("SELECT warehouse_id FROM users WHERE id = :id"),
            {"id": report_dict["user_id"]}
        ).first()
        
        warehouse_id = user_warehouse[0] if user_warehouse else None
        
        if report_dict["equipment_id"] and report_dict["serial_number_id"]:
            # Возвращаем серийный номер
            db.execute(
                text("UPDATE serial_numbers SET status = 'available' WHERE id = :id"),
                {"id": report_dict["serial_number_id"]}
            )
            
            # Возвращаем на склад если есть
            if warehouse_id:
                db.execute(
                    text("UPDATE serial_numbers SET warehouse_id = :wid WHERE id = :id"),
                    {"wid": warehouse_id, "id": report_dict["serial_number_id"]}
                )
        
        elif report_dict["material_id"] and warehouse_id:
            # Возвращаем материал
            db.execute(
                text("""
                    INSERT INTO material_stock (material_id, warehouse_id, quantity)
                    VALUES (:material_id, :warehouse_id, :qty)
                    ON CONFLICT (material_id, warehouse_id) 
                    DO UPDATE SET quantity = material_stock.quantity + :qty
                """),
                {
                    "material_id": report_dict["material_id"],
                    "warehouse_id": warehouse_id,
                    "qty": report_dict["quantity"]
                }
            )
        
        # Обновляем статус отчета
        db.execute(
            text("""
                UPDATE work_reports 
                SET status = 'cancelled', 
                    cancelled_at = now(), 
                    cancelled_by = :user_id,
                    cancel_reason = :reason
                WHERE id = :id
            """),
            {"id": report_id, "user_id": user["id"], "reason": reason}
        )
        
        db.commit()
        
        return {
            "id": report_id,
            "status": "cancelled",
            "message": "Report cancelled and items returned to warehouse"
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error cancelling report: {str(e)}")


@router.delete("/{report_id}")
def delete_report(
    report_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Удалить отчет (только admin, только отмененные)"""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Проверяем статус
    report = db.execute(
        text("SELECT status FROM work_reports WHERE id = :id"),
        {"id": report_id}
    ).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    if report[0] != 'cancelled':
        raise HTTPException(status_code=400, detail="Only cancelled reports can be deleted")
    
    db.execute(
        text("DELETE FROM work_reports WHERE id = :id"),
        {"id": report_id}
    )
    db.commit()
    
    return {"deleted": True, "id": report_id}


# ==================== ЭКСПОРТ ====================

@router.get("/export/csv")
def export_reports_csv(
    user_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Экспорт отчетов в CSV"""
    # Монтажники могут экспортировать только свои отчеты
    if current_user["role"] == "technician":
        user_id = current_user["id"]
    
    reports = list_reports(user_id, date_from, date_to, None, None, db, current_user)
    
    # Создаем CSV
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Заголовок
    headers = [
        "ID", "Дата работы", "Монтажник", "Объект", "Адрес",
        "Тип", "Наименование", "Номер материала", "Серийный номер", 
        "Количество", "Примечание", "Статус", "Создан"
    ]
    writer.writerow(headers)
    
    # Данные
    for r in reports:
        if r["equipment_id"]:
            item_type = "Оборудование"
            name = r["equipment_name"]
            material_num = r["equipment_material_number"]
            serial = r["serial_number"]
        else:
            item_type = "Материал"
            name = r["material_name"]
            material_num = r["material_material_number"]
            serial = "-"
        
        status_labels = {
            "submitted": "Отправлен",
            "approved": "Утвержден",
            "cancelled": "Отменен"
        }
        
        row = [
            r["id"],
            r["work_date"][:19] if r["work_date"] else "",
            r["username"],
            r["object_name"] or "",
            r["object_address"] or "",
            item_type,
            name,
            material_num,
            serial,
            r["quantity"],
            r["notes"] or "",
            status_labels.get(r["status"], r["status"]),
            r["created_at"][:19] if r["created_at"] else ""
        ]
        writer.writerow(row)
    
    output.seek(0)
    
    # Формируем имя файла
    filename_parts = ["reports"]
    if user_id:
        username = db.execute(
            text("SELECT username FROM users WHERE id = :id"),
            {"id": user_id}
        ).scalar()
        filename_parts.append(username)
    filename = "_".join(filename_parts) + ".csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/excel")
def export_reports_excel(
    user_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Экспорт отчетов в Excel (xlsx)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not installed. Run: pip install openpyxl")
    
    # Монтажники могут экспортировать только свои отчеты
    if current_user["role"] == "technician":
        user_id = current_user["id"]
    
    reports = list_reports(user_id, date_from, date_to, None, None, db, current_user)
    
    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчеты"
    
    # Стили
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = [
        "ID", "Дата работы", "Монтажник", "Объект", "Адрес",
        "Тип", "Наименование", "Номер материала", "Серийный номер",
        "Количество", "Примечание", "Статус", "Создан"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Данные
    status_labels = {
        "submitted": "Отправлен",
        "approved": "Утвержден",
        "cancelled": "Отменен"
    }
    
    for row_idx, r in enumerate(reports, 2):
        if r["equipment_id"]:
            item_type = "Оборудование"
            name = r["equipment_name"]
            material_num = r["equipment_material_number"]
            serial = r["serial_number"]
        else:
            item_type = "Материал"
            name = r["material_name"]
            material_num = r["material_material_number"]
            serial = "-"
        
        row_data = [
            r["id"],
            r["work_date"][:10] if r["work_date"] else "",
            r["username"],
            r["object_name"] or "",
            r["object_address"] or "",
            item_type,
            name,
            material_num,
            serial,
            r["quantity"],
            r["notes"] or "",
            status_labels.get(r["status"], r["status"]),
            r["created_at"][:19] if r["created_at"] else ""
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col in [1, 10]:  # Числовые колонки
                cell.alignment = Alignment(horizontal='center')
    
    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Сохраняем в буфер
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Формируем имя файла
    filename_parts = ["reports"]
    if user_id:
        username = db.execute(
            text("SELECT username FROM users WHERE id = :id"),
            {"id": user_id}
        ).scalar()
        filename_parts.append(username)
    filename = "_".join(filename_parts) + ".xlsx"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== СТАТИСТИКА ====================

@router.get("/stats/summary")
def get_stats_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Сводная статистика по отчетам"""
    # Фильтр по датам
    date_filter = ""
    params = {}
    
    if date_from:
        date_filter += " AND work_date >= :date_from"
        params["date_from"] = date_from
    if date_to:
        date_filter += " AND work_date <= :date_to"
        params["date_to"] = date_to
    
    # Общее количество отчетов
    total_query = f"SELECT COUNT(*) FROM work_reports WHERE status != 'cancelled' {date_filter}"
    total = db.execute(text(total_query), params).scalar()
    
    # По монтажникам
    by_user_query = f"""
        SELECT u.username, COUNT(*) as count
        FROM work_reports wr
        JOIN users u ON wr.user_id = u.id
        WHERE wr.status != 'cancelled' {date_filter}
        GROUP BY u.id, u.username
        ORDER BY count DESC
    """
    by_user = db.execute(text(by_user_query), params).fetchall()
    
    # По типам (оборудование/материалы)
    by_type_query = f"""
        SELECT 
            CASE WHEN equipment_id IS NOT NULL THEN 'equipment' ELSE 'material' END as type,
            COUNT(*) as count
        FROM work_reports
        WHERE status != 'cancelled' {date_filter}
        GROUP BY type
    """
    by_type = db.execute(text(by_type_query), params).fetchall()
    
    # По объектам
    by_object_query = f"""
        SELECT object_name, COUNT(*) as count
        FROM work_reports
        WHERE status != 'cancelled' AND object_name IS NOT NULL {date_filter}
        GROUP BY object_name
        ORDER BY count DESC
        LIMIT 10
    """
    by_object = db.execute(text(by_object_query), params).fetchall()
    
    return {
        "total_reports": total,
        "by_user": [{"username": row[0], "count": row[1]} for row in by_user],
        "by_type": [{"type": row[0], "count": row[1]} for row in by_type],
        "by_object": [{"object_name": row[0], "count": row[1]} for row in by_object]
    }